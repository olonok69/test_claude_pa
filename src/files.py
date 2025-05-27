import os
import pandas as pd
from openpyxl import load_workbook
import xlrd
import base64
import io
import re
import uuid
import datetime
import pickle
from IPython.display import HTML, display
from PIL import Image
from langchain_core.documents import Document
from langchain_core.messages import HumanMessage
from src.prompts import pront_multi_system


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def get_sheetnames_xls(filepath):
    xls = xlrd.open_workbook(filepath, on_demand=True)
    return xls.sheet_names()


def load_sheet(filepath, sheet, extension):
    """
    Read sheet from excel or csv. Return Pandas Dataframe
    """
    if extension in ["xlsx", "xls"]:
        df = pd.read_excel(filepath, sheet_name=sheet)
    elif extension == "csv":
        df = pd.read_csv(filepath)

    return df


# Pass the audio data to an encoding function.
def encode_audio(audio_file):
    with open(audio_file, "rb") as f:
        encoded_content = base64.b64encode(f.read())
    return encoded_content


def create_folders(root_dir: str):
    """
    create folders for the project
    """
    OUT_FOLDER = os.path.join(root_dir, "out")
    TMP_FOLDER = os.path.join(root_dir, "tmp")
    ANSWERS_DIR = os.path.join(root_dir, "answers")
    DICTS_DIR = os.path.join(root_dir, "prompts", "dicts")
    PROMPTS_DIR = os.path.join(root_dir, "prompts", "table")
    os.makedirs(TMP_FOLDER, exist_ok=True)
    os.makedirs(ANSWERS_DIR, exist_ok=True)
    os.makedirs(OUT_FOLDER, exist_ok=True)
    os.makedirs(PROMPTS_DIR, exist_ok=True)
    os.makedirs(DICTS_DIR, exist_ok=True)
    return OUT_FOLDER, TMP_FOLDER, ANSWERS_DIR, PROMPTS_DIR, DICTS_DIR


def open_table_pdfs(answers_dir: str):
    """
    open or create table pdf filenames  and restore backup if exists
    params:
        answers_dir: str - directory where the table is stored
    """
    fname = os.path.join(answers_dir, "pdfs.csv")
    fname2 = os.path.join(answers_dir, "pdfs_backup.csv")
    # Open Table
    if os.path.isfile(fname):
        df = pd.read_csv(fname)
        df.to_csv(fname2, index=False)
    else:
        df = pd.DataFrame(columns=["id", "filename", "timestamp"])
        df.to_csv(fname, index=False)
    return fname, fname2, df


def open_table_htmls(answers_dir: str):
    """
    open or create table withs answers and restore backup if exists
    params:
        answers_dir: str - directory where the table is stored
    """
    fname = os.path.join(answers_dir, "htmls.csv")
    fname2 = os.path.join(answers_dir, "htmls_backup.csv")
    # Open Table
    if os.path.isfile(fname):
        df = pd.read_csv(fname)
        df.to_csv(fname2, index=False)
    else:
        df = pd.DataFrame(columns=["id", "url", "quarter", "year", "timestamp"])
        df.to_csv(fname, index=False)
    return fname, fname2, df


def create_urls_dataframe(answers_dir: str):
    """
    open or create table url and content for
    params:
        answers_dir: str - directory where the table is stored
    """
    fname = os.path.join(answers_dir, "htmls.csv")
    fname2 = os.path.join(answers_dir, "htmls_backup.csv")

    url_template1 = "https://nvidianews.nvidia.com/news/nvidia-announces-financial-results-for-{quarter}-quarter-fiscal-{year}"
    url_template2 = "https://nvidianews.nvidia.com/news/nvidia-announces-financial-results-for-{quarter}-quarter-and-fiscal-{year}"
    dataframes = []
    for quarter in ["first", "second", "third", "fourth"]:
        for year in range(2020, 2025):
            args = {"quarter": quarter, "year": str(year)}
            if quarter == "fourth":
                url = url_template2.format(**args)

            else:
                url = url_template1.format(**args)

            p_dict = {}
            p_dict["id"] = str(uuid.uuid4())
            p_dict["url"] = url
            p_dict["quarter"] = quarter
            p_dict["year"] = year
            p_dict["timestamp"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            temp = pd.DataFrame(data=p_dict, index=[0])
            dataframes.append(temp)
    df = pd.concat(dataframes, ignore_index=True)

    df.to_csv(fname, index=False)

    return fname, fname2, df


def save_parsed_documents(fname: str, data: dict):
    """
    save parsed documents to a pickle file
    """

    with open(fname, "wb") as f:
        pickle.dump(data, f)
    return fname


def load_parsed_documents(fname: str):
    """
    load parsed documents from a pickle file
    """

    with open(fname, "rb") as f:
        data = pickle.load(f)
    return data


def plt_img_base64(img_base64):
    """Display base64 encoded string as image"""
    # Create an HTML img tag with the base64 string as the source
    image_html = f'<img src="data:image/jpeg;base64,{img_base64}" />'
    # Display the image by rendering the HTML
    display(HTML(image_html))


def looks_like_base64(sb):
    """Check if the string looks like base64"""
    return re.match("^[A-Za-z0-9+/]+[=]{0,2}$", sb) is not None


def is_image_data(b64data):
    """
    Check if the base64 data is an image by looking at the start of the data
    """
    image_signatures = {
        b"\xff\xd8\xff": "jpg",
        b"\x89\x50\x4e\x47\x0d\x0a\x1a\x0a": "png",
        b"\x47\x49\x46\x38": "gif",
        b"\x52\x49\x46\x46": "webp",
    }
    try:
        header = base64.b64decode(b64data)[:8]  # Decode and get the first 8 bytes
        for sig, format in image_signatures.items():
            if header.startswith(sig):
                return True
        return False
    except Exception:
        return False


def resize_base64_image(base64_string, size=(128, 128)):
    """
    Resize an image encoded as a Base64 string
    """
    # Decode the Base64 string
    img_data = base64.b64decode(base64_string)
    img = Image.open(io.BytesIO(img_data))

    # Resize the image
    resized_img = img.resize(size, Image.LANCZOS)

    # Save the resized image to a bytes buffer
    buffered = io.BytesIO()
    resized_img.save(buffered, format=img.format)

    # Encode the resized image to Base64
    return base64.b64encode(buffered.getvalue()).decode("utf-8")


def split_image_text_types(docs):
    """
    Split base64-encoded images and texts
    """
    b64_images = []
    texts = []
    for doc in docs:
        # Check if the document is of type Document and extract page_content if so
        if isinstance(doc, Document):
            doc = doc.page_content
        if looks_like_base64(doc) and is_image_data(doc):
            doc = resize_base64_image(doc, size=(1300, 600))
            b64_images.append(doc)
        else:
            texts.append(doc)
    if len(b64_images) > 0:
        return {"images": b64_images[:1], "texts": []}
    return {"images": b64_images, "texts": texts}


def img_prompt_func(data_dict):
    """
    Join the context into a single string
    """
    formatted_texts = "\n".join(data_dict["context"]["texts"])
    messages = []

    # Adding the text for analysis
    text_message = {
        "type": "text",
        "text": (
            f"{pront_multi_system}"
            f"User-provided question: {data_dict['question']}\n\n"
            "Text and / or tables:\n"
            f"{formatted_texts}"
        ),
    }
    messages.append(text_message)
    # Adding image(s) to the messages if present
    if data_dict["context"]["images"]:
        for image in data_dict["context"]["images"]:
            image_message = {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpg;base64,{image}"},
            }
            messages.append(image_message)
    return [HumanMessage(content=messages)]
